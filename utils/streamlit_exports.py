import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st

def timedelta_to_excel_numeric(td):
    """Konwertuje timedelta na liczbę odpowiadającą ułamkowi dnia w Excel."""
    if isinstance(td, pd.Timedelta):  # Sprawdzamy, czy td jest obiektem typu Timedelta
        return td.total_seconds() / (24 * 3600)  # Liczba dni jako liczba zmiennoprzecinkowa
    elif isinstance(td, (int, float)):  # Jeśli to jest liczba, zwróć ją bez zmian
        return td
    else:
        return 0  # Jeśli jest to inny typ, zwróć 0 (możesz tu dodać inne zachowanie, jeśli chcesz)

def export_dataframe_to_excel(df):
    """
    Eksportuje DataFrame do pliku Excel, który użytkownik może pobrać w aplikacji Streamlit.
    """

    # Tworzenie pliku Excel w pamięci
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active

    # Konwersja kolumn timedelta na ułamek dnia (dla Excela)
    for column in df.columns:
        if df[column].dtype == "timedelta64[ns]":
            df[column] = df[column].apply(
                lambda x: timedelta_to_excel_numeric(x) if pd.notnull(x) else ""
            )

    # Konwersja kolumny 'wykorzystanie_limitu_rocznego_w_%' na liczby
    if 'wykorzystanie_limitu_rocznego_w_%' in df.columns:
        df['wykorzystanie_limitu_rocznego_w_%'] = pd.to_numeric(df['wykorzystanie_limitu_rocznego_w_%'], errors='coerce')
        df['wykorzystanie_limitu_rocznego_w_%'] = df['wykorzystanie_limitu_rocznego_w_%'].round(2)

    # Konwersja DataFrame na wiersze Excela
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:  # Pomijamy nagłówek przy formatowaniu
            continue
        for c_idx, cell in enumerate(row, start=1):
            header_value = ws.cell(1, c_idx).value
            if header_value == 'wykorzystanie_limitu_rocznego_w_%':  # Formatowanie liczby zmiennoprzecinkowej
                ws.cell(r_idx, c_idx).number_format = '0.00'
            elif isinstance(cell, float):  # Jeśli to ułamek czasu, ustawiamy formatowanie
                ws.cell(r_idx, c_idx).number_format = "[h]:mm:ss"
            elif isinstance(cell, (int, float)):  # Inne wartości numeryczne
                ws.cell(r_idx, c_idx).number_format = '0.00'

    wb.save(output)
    output.seek(0)  # Reset strumienia na początek

    # Udostępnienie pliku do pobrania w Streamlit
    st.download_button(
        label="Pobierz plik Excel",
        data=output,
        file_name="result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def export_dataframe_to_csv(df):
    """
    Eksportuje DataFrame do pliku CSV, który użytkownik może pobrać w aplikacji Streamlit.
    """
    # Konwersja kolumny 'wykorzystanie_limitu_rocznego_w_%' na liczby
    df['wykorzystanie_limitu_rocznego_w_%'] = pd.to_numeric(df['wykorzystanie_limitu_rocznego_w_%'], errors='coerce')

    # Ustawienie zaokrąglenia do dwóch miejsc po przecinku w kolumnie procentowej
    df['wykorzystanie_limitu_rocznego_w_%'] = df['wykorzystanie_limitu_rocznego_w_%'].round(2)

    # Konwersja DataFrame do CSV w pamięci
    csv_data = df.to_csv(index=False, sep=";").encode("utf-8")

    # Udostępnienie pliku do pobrania w Streamlit
    st.download_button(
        label="Pobierz plik CSV",
        data=csv_data,
        file_name="result.csv",
        mime="text/csv",
    )
